"""
Report generation services for analytics.
"""

import os
import io
import base64
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Any, Optional

from django.conf import settings
from django.utils import timezone
from django.template.loader import render_to_string
from django.contrib.auth import get_user_model

# PDF generation
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie

# Excel generation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference

# Charts
import matplotlib
matplotlib.use('Agg')  # Use non-GUI backend
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd

from .models import Report, ReportGeneration, BusinessMetric
from .services import MetricsCalculator

User = get_user_model()


class ReportGenerator:
    """Main report generation service."""
    
    def __init__(self, report: Report, generation: ReportGeneration):
        self.report = report
        self.generation = generation
        self.data = {}
        
        # Create reports directory if it doesn't exist
        self.reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate(self) -> str:
        """Generate report and return file path."""
        try:
            self.generation.status = 'processing'
            self.generation.started_at = timezone.now()
            self.generation.save()
            
            # Collect data
            self._collect_data()
            
            # Generate report based on format
            if self.report.output_format == 'pdf':
                file_path = self._generate_pdf()
            elif self.report.output_format == 'excel':
                file_path = self._generate_excel()
            elif self.report.output_format == 'csv':
                file_path = self._generate_csv()
            else:
                raise ValueError(f"Unsupported format: {self.report.output_format}")
            
            # Update generation record
            self.generation.status = 'completed'
            self.generation.completed_at = timezone.now()
            self.generation.processing_time = self.generation.completed_at - self.generation.started_at
            self.generation.file_path = file_path
            self.generation.file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            self.generation.save()
            
            return file_path
            
        except Exception as e:
            self.generation.status = 'failed'
            self.generation.error_message = str(e)
            self.generation.save()
            raise
    
    def _collect_data(self):
        """Collect data for the report."""
        end_date = self.report.date_range_end or timezone.now().date()
        start_date = self.report.date_range_start or (end_date - timedelta(days=30))
        
        if self.report.report_type == 'daily':
            self.data = self._collect_daily_data(end_date)
        elif self.report.report_type == 'weekly':
            self.data = self._collect_weekly_data(start_date, end_date)
        elif self.report.report_type == 'monthly':
            self.data = self._collect_monthly_data(start_date, end_date)
        elif self.report.report_type == 'quarterly':
            self.data = self._collect_quarterly_data(start_date, end_date)
        elif self.report.report_type == 'yearly':
            self.data = self._collect_yearly_data(start_date, end_date)
        else:
            self.data = self._collect_custom_data(start_date, end_date)
    
    def _collect_daily_data(self, date):
        """Collect data for daily report."""
        return {
            'report_date': date,
            'kpis': MetricsCalculator.calculate_kpi_metrics(date, date),
            'financial': MetricsCalculator.calculate_financial_metrics('day'),
            'user_behavior': MetricsCalculator.calculate_user_behavior_metrics(1)
        }
    
    def _collect_weekly_data(self, start_date, end_date):
        """Collect data for weekly report."""
        return {
            'period': {'start': start_date, 'end': end_date},
            'kpis': MetricsCalculator.calculate_kpi_metrics(start_date, end_date),
            'financial': MetricsCalculator.calculate_financial_metrics('week'),
            'user_behavior': MetricsCalculator.calculate_user_behavior_metrics(7)
        }
    
    def _collect_monthly_data(self, start_date, end_date):
        """Collect data for monthly report."""
        return {
            'period': {'start': start_date, 'end': end_date},
            'kpis': MetricsCalculator.calculate_kpi_metrics(start_date, end_date),
            'financial': MetricsCalculator.calculate_financial_metrics('month'),
            'user_behavior': MetricsCalculator.calculate_user_behavior_metrics(30)
        }
    
    def _collect_quarterly_data(self, start_date, end_date):
        """Collect data for quarterly report."""
        return {
            'period': {'start': start_date, 'end': end_date},
            'kpis': MetricsCalculator.calculate_kpi_metrics(start_date, end_date),
            'financial': MetricsCalculator.calculate_financial_metrics('year'),
            'user_behavior': MetricsCalculator.calculate_user_behavior_metrics(90)
        }
    
    def _collect_yearly_data(self, start_date, end_date):
        """Collect data for yearly report."""
        return {
            'period': {'start': start_date, 'end': end_date},
            'kpis': MetricsCalculator.calculate_kpi_metrics(start_date, end_date),
            'financial': MetricsCalculator.calculate_financial_metrics('year'),
            'user_behavior': MetricsCalculator.calculate_user_behavior_metrics(365)
        }
    
    def _collect_custom_data(self, start_date, end_date):
        """Collect data for custom report."""
        return {
            'period': {'start': start_date, 'end': end_date},
            'kpis': MetricsCalculator.calculate_kpi_metrics(start_date, end_date),
            'financial': MetricsCalculator.calculate_financial_metrics('month'),
            'user_behavior': MetricsCalculator.calculate_user_behavior_metrics(30)
        }
    
    def _generate_pdf(self) -> str:
        """Generate PDF report."""
        filename = f"{self.report.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = os.path.join(self.reports_dir, filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#2c3e50')
        )
        story.append(Paragraph(self.report.name, title_style))
        story.append(Spacer(1, 12))
        
        # Report period
        period_text = f"Período: {self.data.get('period', {}).get('start', '')} - {self.data.get('period', {}).get('end', '')}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Resumo Executivo", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        kpis = self.data.get('kpis', {})
        if kpis:
            # KPI Table
            kpi_data = [
                ['Métrica', 'Valor'],
                ['Usuários Ativos', str(kpis.get('users', {}).get('active_users', 0))],
                ['Total de Reservas', str(kpis.get('bookings', {}).get('total_bookings', 0))],
                ['Receita Total', f"€{kpis.get('financial', {}).get('total_revenue', 0):.2f}"],
                ['Braiders Ativos', str(kpis.get('braiders', {}).get('active_braiders', 0))],
            ]
            
            kpi_table = Table(kpi_data)
            kpi_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(kpi_table)
            story.append(Spacer(1, 20))
        
        # Financial Section
        story.append(Paragraph("Análise Financeira", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        financial = self.data.get('financial', {})
        if financial:
            financial_text = f"""
            Receita Bruta: €{financial.get('totals', {}).get('gross_revenue', 0):.2f}<br/>
            Taxas da Plataforma: €{financial.get('totals', {}).get('platform_fees', 0):.2f}<br/>
            Receita Líquida: €{financial.get('totals', {}).get('net_revenue', 0):.2f}<br/>
            Taxa de Reembolso: {financial.get('totals', {}).get('refund_rate', 0):.1f}%
            """
            story.append(Paragraph(financial_text, styles['Normal']))
            story.append(Spacer(1, 20))
        
        # User Behavior Section
        story.append(Paragraph("Comportamento dos Usuários", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        behavior = self.data.get('user_behavior', {})
        if behavior:
            engagement = behavior.get('engagement', {})
            behavior_text = f"""
            Usuários Ativos: {engagement.get('active_users', 0)}<br/>
            Total de Sessões: {engagement.get('total_sessions', 0)}<br/>
            Atividades por Sessão: {engagement.get('avg_activities_per_session', 0):.1f}<br/>
            Taxa de Retenção: {engagement.get('retention_rate', 0):.1f}%
            """
            story.append(Paragraph(behavior_text, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        return file_path
    
    def _generate_excel(self) -> str:
        """Generate Excel report."""
        filename = f"{self.report.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(self.reports_dir, filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Resumo Executivo"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        
        # Title
        ws_summary['A1'] = self.report.name
        ws_summary['A1'].font = Font(bold=True, size=16)
        
        # Period
        ws_summary['A3'] = "Período:"
        period = self.data.get('period', {})
        ws_summary['B3'] = f"{period.get('start', '')} - {period.get('end', '')}"
        
        # KPIs
        kpis = self.data.get('kpis', {})
        if kpis:
            ws_summary['A5'] = "KPIs Principais"
            ws_summary['A5'].font = header_font
            ws_summary['A5'].fill = header_fill
            
            row = 6
            for section_name, section_data in kpis.items():
                if isinstance(section_data, dict):
                    ws_summary[f'A{row}'] = section_name.title()
                    ws_summary[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    for key, value in section_data.items():
                        ws_summary[f'B{row}'] = key.replace('_', ' ').title()
                        ws_summary[f'C{row}'] = value
                        row += 1
                    row += 1
        
        # Financial sheet
        ws_financial = wb.create_sheet("Financeiro")
        financial = self.data.get('financial', {})
        if financial:
            ws_financial['A1'] = "Análise Financeira"
            ws_financial['A1'].font = Font(bold=True, size=14)
            
            # Totals
            totals = financial.get('totals', {})
            row = 3
            for key, value in totals.items():
                ws_financial[f'A{row}'] = key.replace('_', ' ').title()
                ws_financial[f'B{row}'] = value
                row += 1
            
            # Revenue breakdown
            revenue_breakdown = financial.get('revenue_breakdown', [])
            if revenue_breakdown:
                ws_financial['A10'] = "Breakdown por Tipo"
                ws_financial['A10'].font = Font(bold=True)
                
                # Headers
                ws_financial['A11'] = "Tipo"
                ws_financial['B11'] = "Total"
                ws_financial['C11'] = "Quantidade"
                
                row = 12
                for item in revenue_breakdown:
                    ws_financial[f'A{row}'] = item.get('type', '')
                    ws_financial[f'B{row}'] = item.get('total', 0)
                    ws_financial[f'C{row}'] = item.get('count', 0)
                    row += 1
        
        # Save workbook
        wb.save(file_path)
        return file_path
    
    def _generate_csv(self) -> str:
        """Generate CSV report."""
        filename = f"{self.report.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        file_path = os.path.join(self.reports_dir, filename)
        
        # Create DataFrame from report data
        data_rows = []
        
        # Add KPIs
        kpis = self.data.get('kpis', {})
        if kpis:
            for section_name, section_data in kpis.items():
                if isinstance(section_data, dict):
                    for key, value in section_data.items():
                        data_rows.append({
                            'Category': section_name,
                            'Metric': key,
                            'Value': value
                        })
        
        # Create DataFrame and save
        df = pd.DataFrame(data_rows)
        df.to_csv(file_path, index=False)
        
        return file_path


class ChartGenerator:
    """Service for generating charts and visualizations."""
    
    @staticmethod
    def generate_revenue_trend_chart(data: List[Dict], save_path: str = None) -> str:
        """Generate revenue trend line chart."""
        plt.figure(figsize=(12, 6))
        
        dates = [item['date'] for item in data]
        revenues = [item['revenue'] for item in data]
        
        plt.plot(dates, revenues, marker='o', linewidth=2, markersize=6)
        plt.title('Tendência de Receita', fontsize=16, fontweight='bold')
        plt.xlabel('Data')
        plt.ylabel('Receita (€)')
        plt.xticks(rotation=45)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            plt.close()
            return save_path
        
        # Return base64 encoded image
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        buffer.seek(0)
        image_base64 = base64.b64encode(buffer.read()).decode()
        plt.close()
        
        return f"data:image/png;base64,{image_base64}"
    
    @staticmethod
    def generate_payment_methods_pie_chart(data: List[Dict], save_path: str = None) -> str:
        """Generate payment methods pie chart."""
        plt.figure(figsize=(10, 8))
        
        labels = [item['method'] for item in data]
        sizes = [item['total'] for item in data]
        
        colors = plt.cm.Set3(range(len(labels)))
        
        plt.pie(sizes, labels=labels, autopct='%1.1f%%', colors=colors, startangle=90)
        plt.title('Métodos de Pagamento', fontsize=16, fontweight='bold')
        plt.axis('equal')
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            plt.close()
            return save_path
        
        # Return base64 encoded image
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        buffer.seek(0)
        image_base64 = base64.b64encode(buffer.read()).decode()
        plt.close()
        
        return f"data:image/png;base64,{image_base64}"
    
    @staticmethod
    def generate_user_activity_heatmap(data: List[Dict], save_path: str = None) -> str:
        """Generate user activity heatmap."""
        plt.figure(figsize=(12, 8))
        
        # Create a matrix for heatmap
        activity_matrix = [[0 for _ in range(24)] for _ in range(7)]
        
        for item in data:
            hour = item.get('hour', 0)
            day = item.get('day_of_week', 0)
            count = item.get('count', 0)
            activity_matrix[day][hour] = count
        
        days = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
        hours = [f'{i:02d}:00' for i in range(24)]
        
        sns.heatmap(
            activity_matrix,
            xticklabels=hours[::2],  # Show every 2 hours
            yticklabels=days,
            cmap='YlOrRd',
            annot=False,
            fmt='d'
        )
        
        plt.title('Atividade dos Usuários por Hora', fontsize=16, fontweight='bold')
        plt.xlabel('Hora do Dia')
        plt.ylabel('Dia da Semana')
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            plt.close()
            return save_path
        
        # Return base64 encoded image
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        buffer.seek(0)
        image_base64 = base64.b64encode(buffer.read()).decode()
        plt.close()
        
        return f"data:image/png;base64,{image_base64}"


# Convenience functions
def generate_report(report_id: str, generated_by: User = None) -> ReportGeneration:
    """Generate a report."""
    report = Report.objects.get(id=report_id)
    
    generation = ReportGeneration.objects.create(
        report=report,
        generated_by=generated_by
    )
    
    generator = ReportGenerator(report, generation)
    file_path = generator.generate()
    
    return generation


def generate_executive_summary_pdf() -> str:
    """Generate executive summary PDF."""
    from .services import DashboardService
    
    data = DashboardService.get_executive_dashboard_data()
    
    filename = f"executive_summary_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    file_path = os.path.join(reports_dir, filename)
    
    # Create PDF
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    story.append(Paragraph("Resumo Executivo - Tuwi Beauty", styles['Title']))
    story.append(Spacer(1, 20))
    
    # Key metrics
    key_metrics = data.get('key_metrics', {})
    if key_metrics:
        story.append(Paragraph("Métricas Principais", styles['Heading1']))
        
        for metric_name, metric_data in key_metrics.items():
            current = metric_data.get('current', 0)
            growth = metric_data.get('growth', 0)
            
            metric_text = f"{metric_name.replace('_', ' ').title()}: {current:,.0f}"
            if growth != 0:
                arrow = "↗️" if growth > 0 else "↘️"
                metric_text += f" ({arrow} {growth:+.1f}%)"
            
            story.append(Paragraph(metric_text, styles['Normal']))
        
        story.append(Spacer(1, 20))
    
    # Alerts
    alerts = data.get('alerts', [])
    if alerts:
        story.append(Paragraph("Alertas Ativos", styles['Heading1']))
        
        for alert in alerts:
            alert_text = f"• {alert['title']} ({alert['severity'].upper()})"
            story.append(Paragraph(alert_text, styles['Normal']))
        
        story.append(Spacer(1, 20))
    
    doc.build(story)
    return file_path